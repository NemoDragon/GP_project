import ast
from typing import Callable

import matplotlib.pyplot as plt
from openpyxl import Workbook
from generator import RandomGPlanguageGenerator
from generator_full import RandomGPlanguageGeneratorFull
from languageGP.interpreter import GplInterpreter
from evaluate_functions import EvaluateFunctions
from languageGP.serialization import TreeDeserializer, TreeSerializer
from node import Node
import random
import copy


class GpApp:
    def __init__(self, pop_size=100,
                 crossover_prob=0.05,
                 mutation_prob=0.02,
                 t_size=5,
                 generations=100,
                 init_max_depth=5,
                 init_max_program_size=10,
                 init_max_block_size=3,
                 max_var_number=20,
                 grow_full_ratio=0.5,
                 choose_roulette=False,
                 evaluation_fn=EvaluateFunctions.evaluate,
                 filename='problem.txt'):
        self.pop_size = pop_size
        self.crossover_prob = crossover_prob
        self.mutation_prob = mutation_prob
        self.t_size = t_size
        self.generations = generations
        self.init_max_depth = init_max_depth
        self.init_max_program_size = init_max_program_size
        self.init_max_block_size = init_max_block_size
        self.max_var_number = max_var_number
        self.grow_full_ratio = grow_full_ratio
        self.choose_roulette = choose_roulette
        self.evaluation_fn = evaluation_fn
        self.filename = filename
        self.best_fitness = []
        self.avg_fitness = []
        self.done_generations = []
        self.best_indiv = None
        self.current_gen = 0

    pop = []

    @staticmethod
    def get_all_mutable_nodes(node, elements):
        if isinstance(node, Node):
            for n in node.children:
                if n.node_type in ['if', 'loop', 'assignment', 'out', 'in']:
                    elements.append((n, node))
                    if n.node_type in ['if', 'loop']:
                        GpApp.get_all_mutable_nodes(n.children[1], elements)

    @staticmethod
    def is_variable_context_left_side(variable_node, parent_node):
        if variable_node.node_type == 'variable':
            if parent_node.node_type in ('in', 'assignment', 'array_assignment'):
                if variable_node == parent_node.children[0]:
                    return True
        return False

    def mutate(self, individual):
        prog_size = random.randint(1, self.init_max_program_size)
        block_size = random.randint(1, self.init_max_block_size)
        depth = random.randint(2, self.init_max_depth)
        vars = random.randint(1, self.max_var_number)
        generator = RandomGPlanguageGenerator(prog_size, block_size, depth, vars)

        mutated = copy.deepcopy(individual)
        selected_node, prev_node = GpApp.get_random_node(mutated, with_previous_node=True)
        # take into account context of variable usage
        if GpApp.is_variable_context_left_side(selected_node, prev_node):
            mutated_node = generator.generate_variable()
        else:
            mutated_node = generator.generate_node(selected_node.node_type)
        selected_node.children = mutated_node.children
        selected_node.value = mutated_node.value
        selected_node.node_type = mutated_node.node_type  # Mutated node may have different type
        return mutated

    # excludes program node
    @staticmethod
    def get_random_node(individual, with_previous_node=False):
        individual_nodes = GpApp.flatten_individual_tree(individual)
        if with_previous_node:
            node = random.choice(individual_nodes[1:])
            idx = individual_nodes.index(node) - 1
            prev = individual_nodes[idx - 1]
            return node, prev
        return random.choice(individual_nodes[1:])

    @staticmethod
    def flatten_individual_tree(individual):
        def flatten(node):
            nodes = [node]  # Include the current node
            for child in node.children:  # Recursively include children
                nodes.extend(flatten(child))
            return nodes

        return flatten(individual)

    @staticmethod
    def get_random_node_of_type(individual, node_types):
        nodes = GpApp.flatten_individual_tree(individual)[1:]
        filtered = list(filter(lambda node: node.node_type in node_types, nodes))
        if len(filtered) == 0:
            return None
        return random.choice(filtered)

    @staticmethod
    def crossover(individual1, individual2):
        indiv1 = copy.deepcopy(individual1)
        indiv2 = copy.deepcopy(individual2)

        statement_types = ['if', 'loop', 'assignment', 'out', 'in']
        terminal_types = ['float', 'int', 'variable']

        indiv1_node, indiv1_node_prev = GpApp.get_random_node(indiv1, with_previous_node=True)
        indiv2_node = None

        if indiv1_node.node_type == 'expression':
            indiv2_node = GpApp.get_random_node_of_type(indiv2, ['expression'] + terminal_types)
        elif indiv1_node.node_type == 'logical_condition':
            indiv2_node = GpApp.get_random_node_of_type(indiv2, ['logical_condition'])
        elif indiv1_node.node_type == 'comparison':
            indiv2_node = GpApp.get_random_node_of_type(indiv2, ['comparison'])
        elif indiv1_node.node_type == 'block':
            indiv2_node = GpApp.get_random_node_of_type(indiv2, ['block'])
        elif indiv1_node.node_type in terminal_types:
            if indiv1_node.node_type == 'variable' and GpApp.is_variable_context_left_side(indiv1_node,
                                                                                           indiv1_node_prev):
                indiv2_node = GpApp.get_random_node_of_type(indiv2, ['variable'])
            else:
                indiv2_node = GpApp.get_random_node_of_type(indiv2, terminal_types)
        elif indiv1_node.node_type in statement_types:
            indiv2_node = GpApp.get_random_node_of_type(indiv2, statement_types)

        if indiv2_node is None:
            indiv1_node = GpApp.get_random_node_of_type(indiv1, statement_types)
            indiv2_node = GpApp.get_random_node_of_type(indiv2, statement_types)

        indiv1_node.node_type, indiv2_node.node_type = indiv2_node.node_type, indiv1_node.node_type
        indiv1_node.value, indiv2_node.value = indiv2_node.value, indiv1_node.value
        indiv1_node.children, indiv2_node.children = indiv2_node.children, indiv1_node.children

        return indiv1, indiv2  # In evolution always use indiv1! indiv2 might be semantically incorrect

    @staticmethod
    def output_expression(node):
        if isinstance(node, Node):
            if node.node_type == 'out':
                return GpApp.output_expression(node.children[0])
            if node.node_type == 'expression':
                return GpApp.output_expression(node.children[0]) + node.value + GpApp.output_expression(
                    node.children[1])
            else:
                return node.value

    @staticmethod
    def traverse_node(parent, output):
        if isinstance(parent, Node):
            if parent.children:
                for s in parent.children:
                    output = GpApp.traverse_node(s, output)
            if parent.node_type == 'out':
                output += GpApp.output_expression(parent) if output == "" else "_" + GpApp.output_expression(parent)
        return output


    # tournament - chooses randomly [t_size] programs from population and checks their fitness values with random
    # chosen program from population, program with the lowest fitness value in the tournament is returned
    def tournament(self, t_size=2) -> Node:
        chosen_programs = random.sample(self.pop, k=t_size)
        best_program = random.choice(self.pop)
        best_fitness = self.evaluate_with_problem_file(self.filename, best_program)
        for x in chosen_programs:
            x_fitness = self.evaluate_with_problem_file(self.filename, x)
            if x_fitness < best_fitness:
                best_fitness = x_fitness
                best_program = x
        return best_program

    def negative_tournament(self, t_size=2) -> Node:
        chosen_programs = random.sample(self.pop, k=t_size)
        worst_program = random.choice(self.pop)
        worst_fitness = self.evaluate_with_problem_file(self.filename, worst_program)
        for x in chosen_programs:
            x_fitness = self.evaluate_with_problem_file(self.filename, x)
            if x_fitness > worst_fitness:
                worst_fitness = x_fitness
                worst_program = x
        return worst_program

    # selekcja ruletkowa - losując sposród wszystkich osobników zwraca najgorszego osobnika
    def negative_roulette(self):
        fitnesses_comulative = []
        control_sum = 0
        for program in self.pop:
            eval = self.evaluate_with_problem_file(self.filename, program)
            control_sum += eval
            fitnesses_comulative.append(control_sum)
        fitness_sum = fitnesses_comulative[-1]
        number = random.uniform(0, fitness_sum)
        for i in range(len(fitnesses_comulative)):
            if fitnesses_comulative[i] >= number:
                return self.pop[i]

    # ta zwraca najlepszego
    def roulette(self):
        fitnesses_comulative = []
        control_sum = 0
        for program in self.pop:
            eval = self.evaluate_with_problem_file(self.filename, program)
            if eval < 0.001:
                eval = 0.001
            control_sum += (1 / eval)
            fitnesses_comulative.append(control_sum)
        fitness_sum = fitnesses_comulative[-1]
        number = random.uniform(0, fitness_sum)
        for i in range(len(fitnesses_comulative)):
            if fitnesses_comulative[i] >= number:
                return self.pop[i]

    # create_random_population - generates population of programs and inserts them into an array
    def create_random_population(self) -> None:
        for i in range(self.pop_size):
            prog_size = random.randint(1, self.init_max_program_size)
            block_size = random.randint(1, self.init_max_block_size)
            max_depth = random.randint(2, self.init_max_depth)
            variables = random.randint(1, self.max_var_number)
            if i < self.grow_full_ratio * self.pop_size:
                generator = RandomGPlanguageGenerator(prog_size, block_size, max_depth, variables)
            else:
                generator = RandomGPlanguageGeneratorFull(prog_size, block_size, max_depth, variables)
            prog = generator.generate_program()
            self.pop.append(prog)

    def print_parameters(self) -> None:
        print("GP: programs")
        print("POPSIZE=" + str(self.pop_size))
        print("TSIZE=" + str(self.t_size))
        print("CROSSOVER_PROB=" + str(self.crossover_prob))
        print("MUTATION_PROB=" + str(self.mutation_prob))
        print("INIT_MAX_DEPTH=" + str(self.init_max_depth))
        print("INIT_MAX_PROGRAM_SIZE=" + str(self.init_max_program_size))
        print("INIT_MAX_BLOCK_SIZE=" + str(self.init_max_block_size))
        print("MAX_VAR_NUMBER=" + str(self.max_var_number))
        print("GROW_FULL_RATIO=" + str(self.grow_full_ratio))
        print("CHOOSE_ROULETTE=" + str(self.choose_roulette))
        print("EVOLUTION_FN=" + str(self.evaluation_fn))
        print("GENERATIONS=" + str(self.generations))
        print("FILENAME=" + str(self.filename))

    def stats(self, gen: int) -> float:
        best_individual = self.pop[0]
        best_fitness = self.evaluate_with_problem_file(self.filename, self.pop[0])
        fitness_sum = 0
        for x in self.pop:
            x_fitness = self.evaluate_with_problem_file(self.filename, x)
            fitness_sum += x_fitness
            if x_fitness < best_fitness:
                best_individual = x
                best_fitness = x_fitness
        self.best_fitness.append(best_fitness)
        self.avg_fitness.append(fitness_sum / self.pop_size)
        self.done_generations.append(gen)
        self.best_indiv = best_individual
        print(self.pop_size)
        print(" Best Individual=\n" + str(best_individual) +
              "Generation=" + str(gen) +
              " Avg Fitness=" + str(fitness_sum / self.pop_size) +
              " Best Fitness=" + str(best_fitness))
        return best_fitness

    def save_data_to_excel(self, filename="data.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "data.xlsx"

        # Zapisanie liczb do wierszy
        for i, gen in enumerate(self.done_generations, start=1):
            ws.cell(row=i, column=1, value=gen)
        for i, avg in enumerate(self.avg_fitness, start=1):
            ws.cell(row=i, column=2, value=avg)
        for i, best in enumerate(self.best_fitness, start=1):
            ws.cell(row=i, column=3, value=best)

        program_filename = filename.replace(".xlsx", '.gpl')

        serializer = TreeSerializer()
        serializer.serialize(self.best_indiv, program_filename)

        # Zapisanie pliku
        wb.save(filename)
        print(f"Saved to file: {filename}")

    def plot_data(self):
        fig = plt.figure()
        ax = fig.add_subplot(1, 1, 1)
        #ax.plot(self.done_generations, self.avg_fitness)
        ax.plot(self.done_generations, self.best_fitness, marker="o", markersize=3)
        ax.set_xlabel('generations')
        ax.set_ylabel('best fitness')
        plt.show()
        fig = plt.figure()
        ax = fig.add_subplot(1, 1, 1)
        # ax.plot(self.done_generations, self.avg_fitness)
        ax.plot(self.done_generations, self.avg_fitness, marker="o", markersize=3)
        ax.set_xlabel('generations')
        ax.set_yscale('log')
        ax.set_ylabel('avg fitness')
        plt.show()

    # evolve
    def evolve(self) -> None:
        self.print_parameters()
        self.stats(0)
        for gen in range(1, self.generations):
            self.current_gen = gen
            for i in range(self.pop_size):
                if random.random() < self.crossover_prob:
                    if self.choose_roulette:
                        parent1, parent2 = self.roulette(), self.roulette()
                    else:
                        parent1, parent2 = self.tournament(self.t_size), self.tournament(self.t_size)
                    new_individual = self.crossover(parent1, parent2)[0]
                else:
                    if self.choose_roulette:
                        parent = self.roulette()
                    else:
                        parent = self.tournament(self.t_size)
                    new_individual = self.mutate(parent)
                if self.choose_roulette:
                    offspring = self.negative_roulette()
                else:
                    offspring = self.negative_tournament(self.t_size)
                self.pop.remove(offspring)
                self.pop.append(new_individual)
            best_f = self.stats(gen)
            if best_f <= 0.00001:
                self.save_data_to_excel(self.filename.replace(".txt", '.xlsx'))
                self.plot_data()
                print('PROBLEM SOLVED')
                return
        self.save_data_to_excel(self.filename.replace(".txt", '.xlsx'))
        self.plot_data()
        print('PROBLEM NOT SOLVED')

    def evaluate_with_problem_file(self, file_name: str, program) -> float:
        with open(file_name, 'r') as f:
            data = f.read().splitlines()
        value = 0
        for d in data:
            inputs, outputs = ast.literal_eval(d.split("] [")[0] + "]"), ast.literal_eval("[" + d.split("] [")[1])
            value += self.evaluation_fn(program, inputs, outputs)
        return value


def main():
    gp = GpApp(pop_size=50,
               crossover_prob=0.6,
               t_size=3,
               generations=200,
               init_max_depth=7,
               init_max_program_size=10,
               init_max_block_size=4,
               max_var_number=20,
               grow_full_ratio=0.5,
               choose_roulette=False,
               evaluation_fn=EvaluateFunctions.evaluate_bool_function,
               filename='test_problems/problem_bool.txt')
    gp.create_random_population()
    gp.evolve()


def crossover_test():
    prog_size = random.randint(1, 4)
    block_size = random.randint(1, 3)
    max_depth = random.randint(2, 9)
    generator = RandomGPlanguageGenerator(prog_size, block_size, max_depth)
    prog1 = generator.generate_program()
    prog2 = generator.generate_program()
    gp = GpApp(pop_size=50,
               crossover_prob=0.1,
               mutation_prob=0.1,
               t_size=5,
               generations=200,
               init_max_depth=7,
               filename='')
    print(prog1)
    print('=====================')
    print(prog2)
    print('=====================')
    mutated1, mutated2 = gp.crossover(prog1, prog2)
    print(mutated1)


def mutate_test():
    prog_size = random.randint(1, 4)
    block_size = random.randint(1, 3)
    max_depth = random.randint(2, 9)
    generator = RandomGPlanguageGenerator(prog_size, block_size, max_depth)
    prog = generator.generate_program()
    gp = GpApp(pop_size=50,
               crossover_prob=0.9,
               mutation_prob=0.1,
               t_size=5,
               generations=200,
               init_max_depth=7,
               filename='')
    print(prog)
    mutated = gp.mutate(prog)
    print('=====================')
    print(mutated)


if __name__ == "__main__":
    main()
    #mutate_test()
    #crossover_test()
